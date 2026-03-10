import re
import pandas as pd
import snowflake.connector
from openpyxl.styles import PatternFill
from openpyxl import load_workbook

# ============================================================
# Snowflake connection (as-is)
# ============================================================
def get_connection():
    conn = snowflake.connector.connect(
        user="monopoly22",
        password="8638569740picklU",
        account="JZBSADH-CG45326",
        warehouse="COMPUTE_WH",
        database="DBT_POC",
        schema="DBT_SCHEMA"
    )
    return conn

# ============================================================
# Config
# ============================================================
INPUT_FILE  = "output_with_sql.csv"
OUTPUT_FILE = "data_comp.xlsx"

# Which side’s headers to show for alignment: 'src' or 'tgt'
PREFER_HEADER       = "src"
SORT_BEFORE_COMPARE = True

# ============================================================
# Helpers
# ============================================================
def normalize_headers(df: pd.DataFrame) -> pd.DataFrame:
    """Uppercase, strip whitespace & zero-width chars from headers only."""
    def clean(s):
        s = str(s).replace("\ufeff", "")
        s = re.sub(r"[\u200B-\u200D\uFEFF]", "", s)
        return s.strip().upper()
    out = df.copy()
    out.columns = [clean(c) for c in out.columns]
    return out

def resolve_query_columns(df: pd.DataFrame):
    """
    Locate SQL columns from generator outputs or your own naming.
    """
    src_candidates = ["SRC_QUERY", "SRC", "SOURCE_QUERY", "SOURCE_SQL", "GENERATED_SRC_SQL"]
    tgt_candidates = ["TARGET_QUERY", "TGT_QUERY", "TGT", "TARGET", "TARGET_SQL", "GENERATED_TGT_SQL"]
    colset = set(df.columns)
    src = next((c for c in src_candidates if c in colset), None)
    tgt = next((c for c in tgt_candidates if c in colset), None)
    return src, tgt

def run_query(conn, sql):
    cur = conn.cursor()
    try:
        cur.execute(sql)
        cols = [c[0] for c in cur.description]
        rows = cur.fetchall()
        return pd.DataFrame(rows, columns=cols)
    finally:
        cur.close()

def sort_all_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Stable sort by all columns (string-tuple fallback)."""
    if df.empty or df.shape[1] == 0:
        return df
    by_cols = list(df.columns)
    try:
        return df.sort_values(by=by_cols, kind="mergesort", na_position="last").reset_index(drop=True)
    except Exception:
        df_str = df.applymap(lambda x: "" if pd.isna(x) else str(x))
        df["_SORT_KEY_"] = list(map(tuple, df_str.values.tolist()))
        out = df.sort_values(by="_SORT_KEY_", kind="mergesort", na_position="last") \
                .drop(columns=["_SORT_KEY_"]).reset_index(drop=True)
        return out

def choose_headers(src_df, tgt_df):
    """
    Decide display headers (SRC or TGT). Align both frames by POSITION to same width.
    """
    if PREFER_HEADER.lower() == "tgt" and len(tgt_df.columns) > 0:
        headers = list(tgt_df.columns)
    else:
        headers = list(src_df.columns) if len(src_df.columns) > 0 else list(tgt_df.columns)
    width = min(len(src_df.columns), len(tgt_df.columns), len(headers)) if headers else 0
    headers = headers[:width]
    src_pos = src_df.iloc[:, :width].copy() if width > 0 else src_df.copy()
    tgt_pos = tgt_df.iloc[:, :width].copy() if width > 0 else tgt_df.copy()
    if width > 0:
        src_pos.columns = headers
        tgt_pos.columns = headers
    return headers, src_pos, tgt_pos

def to_str_df(df: pd.DataFrame, headers):
    """Normalize values to strings (trimmed) for robust comparison."""
    out = df.copy()
    for c in headers:
        out[c] = out[c].apply(lambda v: "" if v is None or (isinstance(v, float) and pd.isna(v)) else str(v).strip())
    return out

# ---------------- Multiset A-B / B-A (duplicate-aware) ----------------
def group_counts(df: pd.DataFrame, headers):
    """Group by all columns and return counts (multiset)."""
    if not headers:
        return pd.DataFrame(columns=["CNT"])
    return df.groupby(headers, dropna=False).size().reset_index(name="CNT")

def counts_diff(a_counts: pd.DataFrame, b_counts: pd.DataFrame, headers):
    # FIX: return properly shaped empty frames so downstream code won't break
    if a_counts.empty and b_counts.empty:
        extra_src = pd.DataFrame(columns=headers + ["CNT_SRC", "CNT_TGT", "EXTRA_IN_SRC"]).iloc[0:0]
        extra_tgt = pd.DataFrame(columns=headers + ["CNT_TGT", "CNT_SRC", "EXTRA_IN_TGT"]).iloc[0:0]
        return extra_src, extra_tgt

    ab = a_counts.merge(b_counts, on=headers, how="left", suffixes=("_SRC", "_TGT"))
    ab["CNT_TGT"] = ab["CNT_TGT"].fillna(0).astype(int)
    ab["EXTRA_IN_SRC"] = (ab["CNT_SRC"] - ab["CNT_TGT"]).clip(lower=0).astype(int)
    extra_src = ab[ab["EXTRA_IN_SRC"] > 0].copy()

    ba = b_counts.merge(a_counts, on=headers, how="left", suffixes=("_TGT", "_SRC"))
    ba["CNT_SRC"] = ba["CNT_SRC"].fillna(0).astype(int)
    ba["EXTRA_IN_TGT"] = (ba["CNT_TGT"] - ba["CNT_SRC"]).clip(lower=0).astype(int)
    extra_tgt = ba[ba["EXTRA_IN_TGT"] > 0].copy()

    return extra_src, extra_tgt

def expand_by_count(df: pd.DataFrame, count_col: str, headers):
    """Materialize rows by repeating each row 'count_col' times (preserve duplicates)."""
    if df.empty:
        return pd.DataFrame(columns=headers)
    df = df.loc[df[count_col] > 0].copy()
    df = df.loc[df.index.repeat(df[count_col])].reset_index(drop=True)
    return df[headers].copy()

# ---------------- Duplicate report ----------------
def build_duplicates_report(df, headers, side_label):
    if df.empty or not headers:
        return pd.DataFrame(columns=["SIDE"] + headers + ["COUNT"])
    g = df.groupby(headers, dropna=False).size().reset_index(name="COUNT")
    g = g[g["COUNT"] > 1].copy()
    if g.empty:
        return pd.DataFrame(columns=["SIDE"] + headers + ["COUNT"])
    g.insert(0, "SIDE", side_label)
    return g

# ============================================================
# ROW_ID pairing inside diff sets (no keys)
# ============================================================
def assign_row_id(df):
    """Assign 1..N ROW_ID for each difference set after sorting."""
    if df.empty:
        return df.copy()
    out = df.copy()
    out.insert(0, "ROW_ID", range(1, len(out) + 1))
    return out

def build_mismatch_from_rowid_merge(merged_df, headers, test_id):
    """
    Build mismatch:
      - If both sides present for ROW_ID -> emit stacked SRC/TGT with per-cell highlight
      - If only one side present -> emit SINGLE row (ONLY_IN_SRC / ONLY_IN_TGT)
    """
    rows = []
    row_no = 0

    for _, rec in merged_df.iterrows():
        src_present = any(pd.notna(rec.get(f"{h}_SRC")) and str(rec.get(f"{h}_SRC")).strip() != "" for h in headers)
        tgt_present = any(pd.notna(rec.get(f"{h}_TGT")) and str(rec.get(f"{h}_TGT")).strip() != "" for h in headers)

        if src_present and not tgt_present:
            row_no += 1
            out = {"TEST_ID": test_id, "Row_number": row_no, "type": "ONLY_IN_SRC"}
            for h in headers:
                out[h] = rec.get(f"{h}_SRC")
            rows.append(out)
            continue

        if tgt_present and not src_present:
            row_no += 1
            out = {"TEST_ID": test_id, "Row_number": row_no, "type": "ONLY_IN_TGT"}
            for h in headers:
                out[h] = rec.get(f"{h}_TGT")
            rows.append(out)
            continue

        if not (src_present or tgt_present):
            continue

        # Both present -> skip if identical across all columns
        identical = True
        for h in headers:
            v_src = "" if pd.isna(rec.get(f"{h}_SRC")) else str(rec.get(f"{h}_SRC"))
            v_tgt = "" if pd.isna(rec.get(f"{h}_TGT")) else str(rec.get(f"{h}_TGT"))
            if v_src != v_tgt:
                identical = False
                break
        if identical:
            continue

        row_no += 1
        src_row = {"TEST_ID": test_id, "Row_number": row_no, "type": "SRC"}
        tgt_row = {"TEST_ID": test_id, "Row_number": row_no, "type": "TGT"}
        for h in headers:
            src_row[h] = rec.get(f"{h}_SRC")
            tgt_row[h] = rec.get(f"{h}_TGT")
        rows.append(src_row)
        rows.append(tgt_row)

    if not rows:
        return pd.DataFrame(columns=["TEST_ID", "Row_number"] + headers + ["type"])

    return pd.DataFrame(rows, columns=["TEST_ID", "Row_number"] + headers + ["type"])

# ============================================================
# Write “vertical sections” sheets (as you already had)
# ============================================================
def collect_per_test(frames_list):
    """
    Accepts a list of DataFrames that include TEST_ID, returns dict: {test_id: concatenated_df}.
    """
    per_test = {}
    for df in frames_list:
        if df is None or df.empty or "TEST_ID" not in df.columns:
            continue
        for tid in df["TEST_ID"].unique():
            sub = df[df["TEST_ID"] == tid].copy()
            if sub.empty:
                continue
            per_test.setdefault(int(tid), [])
            per_test[int(tid)].append(sub)
    for tid in list(per_test.keys()):
        per_test[tid] = pd.concat(per_test[tid], ignore_index=True)
    return per_test

def write_vertical_sections(writer, sheet_name: str, per_test: dict):
    """
    Writes each test's dataframe as a vertical section:
      - Title row: "TEST {id}"
      - Header row: dataframe headers (with TEST_ID removed)
      - Data rows below
    """
    from openpyxl.styles import Alignment, Font, PatternFill
    if not per_test:
        pd.DataFrame(columns=["INFO"]).to_excel(writer, sheet_name=sheet_name, index=False)
        return

    current_row = 1
    start_col = 0  # Column A

    for tid in sorted(per_test.keys()):
        df_block = per_test[tid].copy()
        if "TEST_ID" in df_block.columns:
            df_block = df_block.drop(columns=["TEST_ID"])

        df_block.to_excel(writer, sheet_name=sheet_name, index=False,
                          startrow=current_row + 1, startcol=start_col)

        ws = writer.book[sheet_name]
        ncols = df_block.shape[1]
        title_row = current_row + 0
        header_row = current_row + 1
        data_start_row = current_row + 2
        data_last_row = data_start_row + len(df_block) - 1

        title_cell = ws.cell(row=title_row, column=start_col + 1)
        title_cell.value = f"TEST {tid}"
        title_cell.font = Font(bold=True)
        title_cell.alignment = Alignment(horizontal="left")
        gray_fill = PatternFill(start_color="FFD9D9D9", end_color="FFD9D9D9", fill_type="solid")
        title_cell.fill = gray_fill
        if ncols > 1:
            ws.merge_cells(start_row=title_row, start_column=start_col + 1,
                           end_row=title_row, end_column=start_col + ncols)

        current_row = (data_last_row if len(df_block) > 0 else header_row) + 2

# ============================================================
# POST-PROCESS: Highlight mismatches **in place** on 'mismatch' sheet
#   - Tolerates blank rows after TEST title
#   - Tolerates non-contiguous headers (e.g., blank column before TYPE)
#   - Finds Row_number & type by header text, otherwise sniffs TYPE by values
# ============================================================
def _norm(v): 
    return "" if v is None else str(v).strip()

def _find_header_col_idx(ws, header_row, max_c, targets):
    """Find header column index by text (case-insensitive, spaces/dashes ignored)."""
    for c in range(1, max_c + 1):
        hv = ws.cell(header_row, c).value
        if hv is None:
            continue
        h = str(hv).strip().upper().replace(" ", "").replace("-", "_")
        if h in targets:
            return c
    return None

def _sniff_type_col_idx(ws, data_start, data_end, max_c):
    """If TYPE header absent, sniff the column by values."""
    allowed = {"SRC", "TGT", "ONLY_IN_SRC", "ONLY_IN_TGT", ""}
    best_col, best_score = None, -1
    for c in range(1, max_c + 1):
        total = score = 0
        for r in range(data_start, data_end + 1):
            v = _norm(ws.cell(r, c).value).upper()
            if v != "":
                total += 1
                if v in allowed:
                    score += 1
        if total > 0 and score / total >= 0.8 and score > best_score:
            best_col, best_score = c, score
    return best_col

def highlight_mismatch_in_place(filepath: str, sheet_name: str = "mismatch", fill_hex: str = "FFC7CE"):
    wb = load_workbook(filepath)
    if sheet_name not in wb.sheetnames:
        wb.save(filepath)
        return
    ws = wb[sheet_name]
    max_r, max_c = ws.max_row, ws.max_column
    red = PatternFill(start_color=fill_hex, end_color=fill_hex, fill_type="solid")

    r = 1
    while r <= max_r:
        # Find a TEST title in column A
        title = ws.cell(r, 1).value
        if not (isinstance(title, str) and title.strip().upper().startswith("TEST")):
            r += 1
            continue

        # Find header row = first non-blank after title
        hdr = r + 1
        header_row = None
        while hdr <= max_r:
            next_title = isinstance(ws.cell(hdr, 1).value, str) and \
                         ws.cell(hdr, 1).value.strip().upper().startswith("TEST")
            if next_title:
                break
            if any(_norm(ws.cell(hdr, c).value) != "" for c in range(1, max_c + 1)):
                header_row = hdr
                break
            hdr += 1
        if header_row is None:
            r = hdr + 1
            continue

        # Active header columns (non-empty header cells)
        active_cols = [c for c in range(1, max_c + 1) if _norm(ws.cell(header_row, c).value) != ""]
        if not active_cols:
            r = header_row + 1
            continue
        left_bound, right_bound = min(active_cols), max(active_cols)

        # Data start = first non-blank after header
        data_start = header_row + 1
        while data_start <= max_r:
            next_title = isinstance(ws.cell(data_start, 1).value, str) and \
                         ws.cell(data_start, 1).value.strip().upper().startswith("TEST")
            if next_title:
                break
            if any(_norm(ws.cell(data_start, c).value) != "" for c in active_cols):
                break
            data_start += 1

        # Data end = until next TEST or fully blank row across active cols
        data_end = data_start - 1
        rr = data_start
        while rr <= max_r:
            next_title = isinstance(ws.cell(rr, 1).value, str) and \
                         ws.cell(rr, 1).value.strip().upper().startswith("TEST")
            row_blank  = not any(_norm(ws.cell(rr, c).value) != "" for c in active_cols)
            if next_title or row_blank:
                break
            data_end = rr
            rr += 1
        if data_end < data_start:
            r = rr + 1
            continue

        # Find row_number & type columns
        rownum_col = _find_header_col_idx(ws, header_row, max_c, {"ROW_NUMBER", "ROWNUMBER", "ROW_NUM", "ROWNO"})
        type_col   = _find_header_col_idx(ws, header_row, max_c, {"TYPE"})
        if type_col is None:
            type_col = _sniff_type_col_idx(ws, data_start, data_end, max_c)
            if type_col is None:
                r = rr + 1
                continue
        if rownum_col is None:
            rownum_col = left_bound

        # Data columns strictly between rownum and type (blanks allowed)
        lo, hi = min(rownum_col, type_col), max(rownum_col, type_col)
        data_cols = [c for c in range(lo + 1, hi)]

        # Group by Row_number; collect SRC/TGT rows
        groups = {}
        for sr in range(data_start, data_end + 1):
            rn = _norm(ws.cell(sr, rownum_col).value)
            t  = _norm(ws.cell(sr, type_col).value).upper()
            if rn == "":
                continue
            groups.setdefault(rn, {"SRC": [], "TGT": []})
            if t == "SRC":
                groups[rn]["SRC"].append(sr)
            elif t == "TGT":
                groups[rn]["TGT"].append(sr)

        # Paint mismatches in place
        for rn, sides in groups.items():
            pairs = min(len(sides["SRC"]), len(sides["TGT"]))
            for i in range(pairs):
                rs, rt = sides["SRC"][i], sides["TGT"][i]
                for c in data_cols:
                    vs = _norm(ws.cell(rs, c).value)
                    vt = _norm(ws.cell(rt, c).value)
                    if vs == "" and vt == "":
                        continue
                    if vs != vt:
                        ws.cell(rs, c).fill = red
                        ws.cell(rt, c).fill = red

        # Move to next TEST block
        r = rr + 1

    wb.save(filepath)

# ============================================================
# Main
# ============================================================
def main():
    # 1) Read input CSV
    df_in = pd.read_csv(INPUT_FILE, encoding="utf-8-sig", sep=None, engine="python")
    df_in = normalize_headers(df_in)

    src_col, tgt_col = resolve_query_columns(df_in)
    if not src_col or not tgt_col:
        raise ValueError("CSV must contain source & target SQL columns (e.g. 'SRC_QUERY'/'TARGET_QUERY' or 'GENERATED_SRC_SQL'/'GENERATED_TGT_SQL').")

    conn = get_connection()
    summary_rows = []
    a_minus_b_all = []
    b_minus_a_all = []
    mismatch_all = []

    try:
        for i, row in df_in.iterrows():
            test_id = i + 1
            src_sql = str(row[src_col])
            tgt_sql = str(row[tgt_col])

            status = "PASS"
            error  = ""

            try:
                # 2) Execute SRC/TGT SQL
                src_df = run_query(conn, src_sql)
                tgt_df = run_query(conn, tgt_sql)
                
              


                # 3) Align schemas by position & choose headers
                headers, src_w, tgt_w = choose_headers(src_df, tgt_df)
                if not headers:
                    raise ValueError("No comparable columns after alignment (width = 0).")

                # 4) Sort for presentation
                if SORT_BEFORE_COMPARE:
                    src_w = sort_all_columns(src_w)
                    tgt_w = sort_all_columns(tgt_w)

                # 5) Normalize to string
                src_s = to_str_df(src_w, headers)
                tgt_s = to_str_df(tgt_w, headers)
                
                


                # 6–9) Multiset logic & differences
                a_counts = group_counts(src_s, headers)
                b_counts = group_counts(tgt_s, headers)
                extra_src, extra_tgt = counts_diff(a_counts, b_counts, headers)
                a_minus_b = expand_by_count(extra_src.assign(DIFF=extra_src["EXTRA_IN_SRC"]), "DIFF", headers)
                b_minus_a = expand_by_count(extra_tgt.assign(DIFF=extra_tgt["EXTRA_IN_TGT"]), "DIFF", headers)
                a_minus_b = sort_all_columns(a_minus_b)
                b_minus_a = sort_all_columns(b_minus_a)
                a_minus_b_id = assign_row_id(a_minus_b)
                b_minus_a_id = assign_row_id(b_minus_a)

                if not a_minus_b_id.empty:
                    tmp = a_minus_b_id.copy()
                    tmp.insert(0, "TEST_ID", test_id)
                    a_minus_b_all.append(tmp)
                if not b_minus_a_id.empty:
                    tmp = b_minus_a_id.copy()
                    tmp.insert(0, "TEST_ID", test_id)
                    b_minus_a_all.append(tmp)

                # 10–11) Build mismatch rows
                if a_minus_b_id.empty and b_minus_a_id.empty:
                    status = "PASS"
                    merged = pd.DataFrame()
                else:
                    status = "FAIL"
                    merged = a_minus_b_id.merge(
                        b_minus_a_id,
                        on=["ROW_ID"],
                        how="outer",
                        suffixes=("_SRC", "_TGT"),
                        indicator=True
                    ).sort_values(["ROW_ID"], kind="mergesort").reset_index(drop=True)

                mm_df = build_mismatch_from_rowid_merge(merged, headers, test_id)
                if not mm_df.empty:
                    mismatch_all.append(mm_df)

            except Exception as ex:
                status = "FAIL"
                error = str(ex)

            summary_rows.append([src_sql, tgt_sql, status, error])

    finally:
        try:
            conn.close()
        except Exception:
            pass

    # 12) Write Excel
    summary_df = pd.DataFrame(summary_rows, columns=["SRC_QUERY", "TARGET_QUERY", "STATUS", "ERROR_MESSAGE"])

    with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
        # Summary
        summary_df.to_excel(writer, sheet_name="data_comp", index=False)

        # A_MINUS_B
        if a_minus_b_all:
            per_test_a = collect_per_test(a_minus_b_all)
            write_vertical_sections(writer, "A_MINUS_B", per_test_a)
        else:
            pd.DataFrame(columns=["INFO"]).to_excel(writer, sheet_name="A_MINUS_B", index=False)

        # B_MINUS_A
        if b_minus_a_all:
            per_test_b = collect_per_test(b_minus_a_all)
            write_vertical_sections(writer, "B_MINUS_A", per_test_b)
        else:
            pd.DataFrame(columns=["INFO"]).to_excel(writer, sheet_name="B_MINUS_A", index=False)

        # mismatch (no colors yet; colors will be applied in-place below)
        if mismatch_all:
            per_test_m = collect_per_test(mismatch_all)
            write_vertical_sections(writer, "mismatch", per_test_m)
        else:
            pd.DataFrame(columns=["INFO"]).to_excel(writer, sheet_name="mismatch", index=False)

    # 13) Highlight mismatches IN PLACE on 'mismatch'
    highlight_mismatch_in_place(OUTPUT_FILE, sheet_name="mismatch", fill_hex="FFC7CE")

    print("DONE — data_comp.xlsx created and 'mismatch' colored in place (only differing SRC/TGT cells).")

if __name__ == "__main__":
    main()
